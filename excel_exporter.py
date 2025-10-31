"""
Excel Exporter Module
Handles exporting manufacturer analysis results to Excel with formatting
"""

import pandas as pd
import logging
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelExporter:
    """Exports manufacturer analysis results to formatted Excel files"""
    
    def __init__(self, output_path: Optional[str] = None):
        """
        Initialize ExcelExporter
        
        Args:
            output_path (str, optional): Output file path. Auto-generated if not provided
        """
        if output_path:
            self.output_path = output_path
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_path = f"manufacturer_analysis_{timestamp}.xlsx"
    
    def export(self, df: pd.DataFrame) -> str:
        """
        Export DataFrame to Excel with formatting
        
        Args:
            df (pd.DataFrame): DataFrame to export
            
        Returns:
            str: Path to the exported file
        """
        try:
            logger.info(f"Exporting results to {self.output_path}")
            
            # Write to Excel
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Manufacturer Analysis', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Manufacturer Analysis']
                
                # Apply formatting
                self._format_worksheet(worksheet, df)
            
            logger.info(f"Successfully exported to {self.output_path}")
            return self.output_path
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame):
        """
        Apply formatting to the worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            df (pd.DataFrame): Source DataFrame
        """
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Credibility score fills
        high_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        medium_score_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        low_score_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = border
                
                # Apply conditional formatting for credibility scores
                if df.columns[col_num - 1] == 'Avg_Credibility_Score':
                    try:
                        score = float(cell.value)
                        if score >= 80:
                            cell.fill = high_score_fill
                        elif score >= 60:
                            cell.fill = medium_score_fill
                        else:
                            cell.fill = low_score_fill
                    except (ValueError, TypeError):
                        pass
        
        # Adjust column widths
        column_widths = {
            'ID': 8,
            'MPN': 20,
            'Model_Description': 40,
            'Quantity': 12,
            'Top_Manufacturer': 25,
            'All_Manufacturers': 35,
            'Avg_Credibility_Score': 18,
            'Recommendation': 45,
            'Detailed_Analysis': 50,
            'Additional_Info': 35
        }
        
        for col_num, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            width = column_widths.get(column, 15)
            worksheet.column_dimensions[col_letter].width = width
        
        # Set row heights
        worksheet.row_dimensions[1].height = 30
        for row_num in range(2, len(df) + 2):
            worksheet.row_dimensions[row_num].height = 60
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
    
    def create_summary_sheet(self, df: pd.DataFrame, output_path: Optional[str] = None) -> str:
        """
        Create an Excel file with both detailed and summary sheets
        
        Args:
            df (pd.DataFrame): DataFrame with manufacturer analysis
            output_path (str, optional): Output file path
            
        Returns:
            str: Path to the exported file
        """
        if output_path:
            self.output_path = output_path
        
        try:
            logger.info(f"Creating summary report at {self.output_path}")
            
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                # Write detailed analysis
                df.to_excel(writer, sheet_name='Detailed Analysis', index=False)
                
                # Create summary DataFrame
                summary_df = df[['ID', 'MPN', 'Model_Description', 'Top_Manufacturer', 'Avg_Credibility_Score']].copy()
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format both sheets
                self._format_worksheet(writer.sheets['Detailed Analysis'], df)
                self._format_worksheet(writer.sheets['Summary'], summary_df)
            
            logger.info(f"Successfully created summary report at {self.output_path}")
            return self.output_path
            
        except Exception as e:
            logger.error(f"Error creating summary report: {str(e)}")
            raise

